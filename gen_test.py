
#modules importing
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import unittest

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.use_chromium = True 

service = Service(executable_path="c:\\users\\shaba7\\appdata\\local\\programs\\python\\python39\\msedgedriver.exe",verbose = True)
driver = webdriver.Edge(service=service, options=options) 
from bs4 import BeautifulSoup as bs
import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
import re
import datetime

import dropbox
from tqdm import tqdm
now = datetime.datetime.now()

def connect_dropbox():
     #you shoul create dropbox application and change these 2 variables with yours 
    app_key = ""
    app_secret = ""
    auth_flow = dropbox.DropboxOAuth2FlowNoRedirect(app_key, app_secret)

    auth_url = auth_flow.start()
    
    try:
        auth_code = get_auth_link(auth_url)
    except:
        print()
        print("\33[33m1. Go to: " + auth_url)
        print("2. Click \"Allow\" (you might have to log in first).")
        print("3. Copy the authorization code.\33[0m")
        auth_code = input("\33[4;33mEnter the authorization code here: \33[0m").strip()
    
    try:
        oauth_result = auth_flow.finish(auth_code)
    except Exception as e:
        print('Error: %s' % (e,))
        exit(1)
    with dropbox.Dropbox(oauth2_access_token=oauth_result.access_token) as dbx:
        print("Successfully set up client!")
        print()
        dbxs = dbx
    return dbxs


####### function of making shareable link
def get_drop_link(file_path_code,dbx,status):
    dbx =dbx
    file_path = file_path_code.replace(" ","_")
    category_name = file_path_code.split(' ')[0]
    status = status
    if status == 'NO DATA':
        shareble_link = 'NO DATA'
    else:
        ##########     this path need to be changed to add category state folder path too ########################################################
        shared_link_metadata = dbx.sharing_create_shared_link("/Abdelwahab Ahmed/03 My work/Trade Service/"+category_name+"/Trade_Service_"+file_path+".xlsx")
        shared_link = shared_link_metadata.url
        shareble_link = shared_link
    
    return shareble_link 

#####function of getting authorization link from dropbox
def get_auth_link(auth_url):
    driver = webdriver.Edge(service=service, options=options) 
    driver.get(auth_url)
    try:
        holder = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".credentials-form__fields")))
        try:
            children = holder.find_elements(By.CSS_SELECTOR, "input[type]")
            for i,ch in enumerate(children):
                if ch.get_attribute("type") == 'email':
                    ch.send_keys("") ## add your email of dropbox
                elif ch.get_attribute("type") == 'password':
                    ch.send_keys("") ### add your dropbox password
        except:
            ch_email= driver.find_element(By.NAME, 'login_email').send_keys("")## add your email of dropbox
            ch_pass = driver.find_element(By.NAME, 'login_password').send_keys("")### add your dropbox password
    

        try:
            button = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button .signin-text')))
            button.click()
        except:
            driver.get("https://www.dropbox.com/oauth2/authorize?response_type=code&client_id=jo4g9e15ynnmocd")
            button = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'form button .signin-text')))
            button.click()
    except:
        pass

    try:
        cont_but = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, 'warning-button-continue')))
        cont_but.click() 
    except:
        cont_but = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.app-warning-frame button')))
        for i,cont in enumerate(cont_but):
            if cont.get_attribute('id') == 'warning-button-continue':
                cont.click()

    try:
        allow_but = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.NAME, 'allow_access')))
        allow_but.click()
    except:
        butons = driver.find_elements(By.CSS_SELECTOR, '#buttons button')
        for i,b in enumerate(butons):
            if b.text == 'Allow':
                b.click()

    link_holder = driver.find_element(By.CSS_SELECTOR, '.auth-connect-scoped-frame input.auth-box')
    return link_holder.get_attribute("data-token")

def generate_file(shareable_lnk,diagnostic):
    sh_lnk = shareable_lnk
    df = diagnostic
    category_name = df['category']
    
    suburb = df['suburb']
    
    state = df['state']
    
    code_postcode = df['file path']
    scraped_time = df['scrapped date']
    scraped_rows = df['number of services']
    status = df['status']

    try:
        ##read and write
        wb = load_workbook('Trade_service_generator.xlsx')
        ws = wb.active
        
        ws.append([category_name,suburb,state,postcode,code_postcode,scraped_time,scraped_rows,status,sh_lnk ])
        wb.save(filename='Trade_service_generator.xlsx')
    except:
        headers  = ['category','state','suburb','file path','scrapped date','number of services', 'status','dropbox link']
        workbook_name = 'Trade_service_generator.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append([category_name,suburb,state,postcode,code_postcode,scraped_time,scraped_rows,status,sh_lnk ])
        wb.save(workbook_name)
    return 

dbx = connect_dropbox()
def operate(dbx):
    dbx = dbx

    diagnostic= pd.read_excel("Trade_service_diagnostic.xlsx")
    status = diagnostic['status']
    try:
        generator_file = pd.read_excel("Trade_service_generator.xlsx")
        latest_row = generator_file['path code'].loc[len(generator_file['path code'])-1]
        row_length = len(generator_file)
    except:
        row_length = 0
    print(f"-------------------------------------------- \n>>>> you have {row_length}/{len(status)} shareable link <<<< \n--------------------------------------------")
    if row_length == 0:
            try:
                if status[0] == 'Scrapped Successfully':
                    
                    shareable_lnk = get_drop_link(diagnostic['path code'].loc[0],dbx,diagnostic['status'].loc[0])
                    generate_file(shareable_lnk,diagnostic.loc[0])
                    operate(dbx)
                    
                    
                elif status[0] =='NO DATA':
                    shareable_lnk = 'NO DATA'
                    generate_file(shareable_lnk,diagnostic.loc[0])
                    operate(dbx)

            except Exception as e: 
                print(f"------------------------ \nyou got an exception: '{e}' \nits index is 0 \n------------------------ ")
                
        

    elif row_length > 0:
        for i,x in enumerate(diagnostic['path code']):
            generator_file = pd.read_excel("Trade_service_generator.xlsx")
            latest_row = generator_file['path code'].loc[len(generator_file['path code'])-1]
            if i == len(diagnostic['path code'])-1 :
                break
            if x == latest_row :
                shareable_lnk = get_drop_link(diagnostic['path code'].loc[i+1],dbx,diagnostic['status'].loc[i+1])
                generate_file(shareable_lnk,diagnostic.loc[i+1])
            
            else:
                continue

    for i in tqdm ( range(len(status)), desc=f"generating {len(status) - row_length} files",ascii=False,ncols=75): ### if you found any problem about this line you may check it online or remove it 
        time.sleep(0.01)
    return 

    

operate(dbx)
