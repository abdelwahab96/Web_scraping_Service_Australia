#import modules and packages 
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import unittest

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.use_chromium = True 

service = Service(executable_path="c:\\users\\shaba7\\appdata\\local\\programs\\python\\python39\\msedgedriver.exe",verbose = True)
driver = webdriver.Edge(service=service, options=options) 

from bs4 import BeautifulSoup as bs
import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
import re
import datetime
import dropbox
now = datetime.datetime.now()




def operation(category_name,category_link,index,cat_length):
    try:
        if driver is None:
            driver = webdriver.Edge(service=service, options=options)
    except:
        pass
    ### open categories file
    ## function of getting the category link from the excel
    #category_link , category_name = get_category_link_nd_name() # this function returning category link
    states_lnks = get_states_links(category_link)
    print(f"-------------------------------- \nNow you are counting the number of services for '{category_name}' \nit's the category number {index} / {cat_length} \n--------------------------------")
    data_row= []
    data_row.append(category_name)
    data_row.append(category_link)
    suburbs_all_length = 0
    total_services_states = 0
    premium_services_states =0
    basic_services_states  = 0
    for i,state_link in enumerate(states_lnks):
        # state number 1 QLD
        suburbs_length , suburbs_links = get_suburbs_length_nd_links(state_link)

        total_services_suburb =0
        premium_services_suburb =0
        basic_services_suburb =0
        for ind, each_suburb in enumerate(suburbs_links):
            services_lnks_nd_types = serv_link_nd_type(each_suburb)
            tot, pr, ba = count_service_type(services_lnks_nd_types)
            total_services_suburb += tot
            premium_services_suburb += pr
            basic_services_suburb += ba
            #now u got into one suburb 
            if ind == 0:
                break
        total_services_states += total_services_suburb
        premium_services_states += premium_services_suburb
        basic_services_states += basic_services_suburb
        suburbs_all_length += suburbs_length
        data_row.append(suburbs_length)
        data_row.append(total_services_suburb)
        data_row.append(premium_services_suburb)
        data_row.append(basic_services_suburb)
        
    data_row.append(suburbs_all_length)
    data_row.append(total_services_states)
    data_row.append(premium_services_states)
    data_row.append(basic_services_states)
    full_data_row = data_row
    gen_file(full_data_row)


    return print(f"--------------------------- \nnow you finished {index}/{cat_length} category services counting \n---------------------------")


    
    










############# getting the category link ##################

def get_category_link_nd_name():

    categories = pd.read_excel('categories_links.xlsx')
    try:
        counter_file = pd.read_excel('Trade_service_counter.xlsx')
        num_rows = len(counter_file)
        latest_row = counter_file['Category'].loc[len(counter_file) -1]
    except:
        num_rows = 0
    for i, x in enumerate(categories['name']):
        if num_rows == 0:
            category_name = categories['name'].loc[0]
            category_link =categories['categories_href'].loc[0]
            break
        if x == latest_row:
            category_link = categories['categories_href'].loc[i+1]
            category_name = categories['name'].loc[i+1]
            break
        elif x != latest_row:
            continue
        else:
            print("there is no category with this name!")
    return category_link ,category_name
    


############### get states links >> always 8 ################ 
def get_states_links(categ):
    
    #driver = webdriver.Edge(service=service, options= options)
    driver.get(categ)
    holding = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.box-content')))
    states = driver.find_elements(By.CSS_SELECTOR, '.box-content')
    states_lst = states[0].find_elements(By.CSS_SELECTOR, 'a')
    states_names = [j.get_attribute('href').split('/')[-2] for j in states_lst]
    states_lnks = [i.get_attribute('href') for i in states_lst]
    
    return states_lnks



############# getting the number of suburbs ###############
def get_suburbs_length_nd_links(state_link):
    driver.get(state_link)
    suburb_lst = driver.find_elements(By.CSS_SELECTOR, 'ul.block-list li > a')
    suburbs_lnks = [suburb_lnk.get_attribute('href') for suburb_lnk in suburb_lst ]
    return len(suburb_lst) , suburbs_lnks



##### get next links if it exist #####
def get_next_links(suburb_link):
    lst_service = [suburb_link]
    buttons = driver.find_elements(By.CSS_SELECTOR, '.pagination nav > *')

    while buttons[1].get_attribute('rel') == 'next':
        next_lnk = buttons[1].get_attribute('href')
        lst_service.append(next_lnk)
        driver.get(next_lnk)
        buttons = driver.find_elements(By.CSS_SELECTOR, '.pagination nav > *')
        
    return lst_service



############### function of getting services length and types ########## 
####### take suburb return services and service data type (full/ partial)
def serv_link_nd_type(suburb_link):
    services_lnks_nd_types = []
    driver.get(suburb_link)
    
    try:
        #check if there is next
        driver.find_element(By.CSS_SELECTOR, '.pagination')
        suburbs_lnks = get_next_links(suburb_link)
        
        for lnk in suburbs_lnks:
            driver.get(lnk)
            sevices_lst = driver.find_elements(By.CSS_SELECTOR, '.results-cont > .directory-listing-box')
            for i in sevices_lst:
                data_type = i.get_attribute('class').split(" ")[1]
                service_lnk = i.find_element(By.CSS_SELECTOR, '.listing-info  a').get_attribute('href')
                services_lnks_nd_types.append([service_lnk,data_type])
    except: # there is no next links
        sevices_lst = driver.find_elements(By.CSS_SELECTOR, '.results-cont > .directory-listing-box')
        try:
            for i in sevices_lst:
                data_type = i.get_attribute('class').split(" ")[1]
                service_lnk = i.find_element(By.CSS_SELECTOR, '.listing-info  a').get_attribute('href')
                services_lnks_nd_types.append([service_lnk,data_type])
        except:
            data_type = sevices_lst.get_attribute('class').split(" ")[1]
            service_lnk = sevices_lst.find_element(By.CSS_SELECTOR, '.listing-info  a').get_attribute('href')
            services_lnks_nd_types.append([service_lnk,data_type])
    
    
    
    return services_lnks_nd_types





######## count services types ##############
def count_service_type(services_lnks_nd_types):   
    total_services_type = len(services_lnks_nd_types)
    premium_services_type = 0
    basic_services_type = 0
    for item in services_lnks_nd_types:
        if item[1] == 'premium':
            premium_services_type +=1
            continue
        if item[1] == 'basic':
            basic_services_type += 1
            continue
    
    return total_services_type, premium_services_type, basic_services_type



    ########### function of generating excel file
def gen_file(full_data_row):
    
    try:
        ##read and write
        wb = load_workbook('Trade_service_counter.xlsx')
        ws = wb.active
        ws.append(full_data_row)
        wb.save(filename='Trade_service_counter.xlsx')
    except:
        headers  = ['Category', 'Category Link',
                     'QLD - Suburbs' , 'QLD - Total Data', 'QLD - Premium Data','QLD - Basic Data',
                    'VIC - Suburbs' , 'VIC - Total Data', 'VIC - Premium Data','VIC - Basic Data',
                    'ACT - Suburbs' , 'ACT - Total Data', 'ACT - Premium Data','ACT - Basic Data',
                    'NSW - Suburbs' , 'NSW - Total Data', 'NSW - Premium Data','NSW - Basic Data',
                    'WA - Suburbs' , 'WA - Total Data', 'WA - Premium Data','WA - Basic Data',
                    'TAS - Suburbs' , 'TAS - Total Data', 'TAS - Premium Data','TAS - Basic Data',
                    'SA - Suburbs' , 'SA - Total Data', 'SA - Premium Data','SA - Basic Data',
                    'NT - Suburbs' , 'NT - Total Data', 'NT - Premium Data','NT - Basic Data',
                     'Scrapable Suburbs','Total Services (Both Data)','Total Premium Data','Total Basic Data' ]
        workbook_name = 'Trade_service_counter.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(full_data_row)
        wb.save(workbook_name)
    return





    

def start_scrap():
    df_categories = pd.read_excel('categories_links.xlsx')
    
    
    for i,x in enumerate(df_categories['name']):
        try:
            df_counter = pd.read_excel('Trade_service_counter.xlsx')
            latest_row = df_counter['Category'].loc[len(df_counter)- 1]
            
            
            index = len(df_counter)
            num_rows = len(df_counter)
        except:
            
            num_rows = 0
            index = 1
        

        if num_rows == 0 :
            
            category_name= df_categories['name'].loc[0]
            category_link = df_categories['categories_href'].loc[0]
            operation(category_name, category_link, i+1, len(df_categories))
            start_scrap()
        if num_rows > 0:
            if x == latest_row:
                print(latest_row)
                category_name= df_categories['name'].loc[i+1]
                category_link = df_categories['categories_href'].loc[i+1]
                operation(category_name, category_link, i+2, len(df_categories))
            else:
                continue
        try:
            if len(df_counter) == len(df_categories):
                break
        except:
            pass
    
    return


start_scrap()
try:
    driver.dispose()
except:
    driver.quit()