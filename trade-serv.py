#import modules and packages 
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import unittest

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.use_chromium = True 

service = Service(executable_path="c:\\users\\shaba7\\appdata\\local\\programs\\python\\python39\\msedgedriver.exe",verbose = True)
driver = webdriver.Edge(service=service, options=options) 

from bs4 import BeautifulSoup as bs
import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import xlsxwriter
import re
import datetime
import dropbox
now = datetime.datetime.now()


###### getting the link of the category name   ######################
category_name= 'Abrasive'
def get_category_link(category):
    categories = pd.read_excel('categories_links.xlsx')
    for i, x in enumerate(categories['name']):
        if x == category:
            category_link = categories['categories_href'].loc[i]
        elif x != category:
            continue
        else:
            print("there is no category with this name!")
    return category_link 

########### get the state links  ######
def get_states_links(categ):
    
    #driver = webdriver.Edge(service=service, options= options)
    driver.get(categ)
    holding = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.box-content')))
    states = driver.find_elements(By.CSS_SELECTOR, '.box-content')
    states_lst = states[0].find_elements(By.CSS_SELECTOR, 'a')
    states_names = [j.get_attribute('href').split('/')[-2] for j in states_lst]
    states_lnks = [i.get_attribute('href') for i in states_lst]
    
    return states_lnks, states_names

#get the suburb link
def get_suburbs(state_link):
    
    driver.get(state_link)
    suburb_lst = driver.find_elements(By.CSS_SELECTOR, 'ul.block-list li > a')
    suburbs_lnks = [suburb_lnk.get_attribute('href') for suburb_lnk in suburb_lst ]
    suburb_names = [name.text for name in suburb_lst]
    return suburbs_lnks ,suburb_names




####### take suburb return services and service data type (full/ partial)
def serv_link_nd_type(suburb_link):
    services_lnks_nd_types = []
    driver.get(suburb_link)
    
    try:
        #check if there is next
        driver.find_element(By.CSS_SELECTOR, '.pagination')
        services_lnks_nd_types = get_next_links(suburb_link)
        
    except: # there is no next links
        sevices_lst = driver.find_elements(By.CSS_SELECTOR, '.results-cont > .directory-listing-box')
        try:
            for i in sevices_lst:
                data_type = i.get_attribute('class').split(" ")[1]
                service_lnk = i.find_element(By.CSS_SELECTOR, '.listing-info  a').get_attribute('href')
                services_lnks_nd_types.append([service_lnk,data_type])
        except:
            data_type = sevices_lst.get_attribute('class').split(" ")[1]
            service_lnk = sevices_lst.find_element(By.CSS_SELECTOR, '.listing-info  a').get_attribute('href')
            services_lnks_nd_types.append([service_lnk,data_type])
    
    
    
    return services_lnks_nd_types

####### function of getting links of next button #######
def get_next_links(suburb_link):
    #driver.get(suburb_link)
    lst_service = [suburb_link]
    buttons = driver.find_elements(By.CSS_SELECTOR, '.pagination nav > *')
    services_lnks_nd_types = []
    
    #next_lnk = buttons[1].get_attribute('href')
    #lst_service.append(next_lnk)#disabled
    next_class = buttons[1].get_attribute('class')
    
    while next_class == '':
        sevices_lst = driver.find_elements(By.CSS_SELECTOR, '.results-cont > .directory-listing-box')
        buttons = driver.find_elements(By.CSS_SELECTOR, '.pagination nav > *')
        
        for i in sevices_lst:
            data_type = i.get_attribute('class').split(" ")[1]
            service_lnk = i.find_element(By.CSS_SELECTOR, '.listing-info  a').get_attribute('href')
            services_lnks_nd_types.append([service_lnk,data_type])
        if buttons[1].get_attribute('class') == 'disabled':
            next_class = 'disabled'
            break
        else:

            buttons = driver.find_elements(By.CSS_SELECTOR, '.pagination nav > *')
            next_lnk = buttons[1].get_attribute('href')
            driver.get(next_lnk)
        
    return services_lnks_nd_types

### scraping function #######
def scrap(service_link,category_name,state,suburb ,service_data_type ):
    service = service_link
    data_type = service_data_type
    driver.get(service)
    title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1')))
    try:
        serv_title = driver.find_element(By.CSS_SELECTOR, 'h1').text
    except:
        serv_title = 'NA'
    
    if service_data_type == 'premium':
        try:
            social_lst = driver.find_elements(By.CSS_SELECTOR, '.social-list')
            address_lst = social_lst[0]
            
            try:
                location = address_lst.text.split('\n')[1]
            except:
                location = address_lst.text

            try:
                social_media_lst = social_lst[1].find_elements(By.CSS_SELECTOR, 'li a')
                try:
                    if len(social_media_lst) >1:
                        for i,x in enumerate(social_media_lst):
                            if i == 0:
                                item = x.get_attribute('href')
                                website = item
                                continue
                            item = x.get_attribute('href')
                            if item.split('.')[1] == 'facebook':
                                facebook = item 
                            
                            
                        twt = 'NA'
                        linkedin = 'NA'
                        insta = 'NA'
                        tktok = 'NA'
                
                                

                    elif len(social_media_lst) ==1 :
                        for i in social_media_lst:
                            item = i.get_attribute('href')
                            if item.split('.')[1] == 'facebook':
                                facebook = item 
                            elif item.split('.')[1] == 'instagram':
                                insta = item
                            elif item.split('.')[1] == 'linkedin':
                                linkedin = item
                            elif item.split('.')[1] == 'twitter':
                                twt = item
                            elif item.split('.')[1] == 'tiktok':
                                tktok = item
                            else:
                                website = item
                                twt = 'NA'
                                linkedin = 'NA'
                                facebook = 'NA'
                                insta = 'NA'
                                tktok = 'NA'
                except:
                    
                    for i in social_media_lst:
                        item = i.get_attribute('href')
                        if item.split('.')[1] == 'facebook':
                            facebook = item 
                        elif item.split('.')[1] == 'instagram':
                            insta = item
                        elif item.split('.')[1] == 'linkedin':
                            linkedin = item
                        elif item.split('.')[1] == 'twitter':
                            twt = item
                        elif item.split('.')[1] == 'tiktok':
                            tktok = item
                        else:
                            website = item
                            twt = 'NA'
                            linkedin = 'NA'
                            facebook = 'NA'
                            insta = 'NA'
                            tktok = 'NA'
 
            
            except:
                facebook = 'NA'
                insta = 'NA'
                website = 'NA'
            
        except:
            location = 'NA'
            facebook = 'NA'
            insta = 'NA'
            website = 'NA'

        ### about us and provided services
        provided_services = []
        services_obj = driver.find_elements(By.CSS_SELECTOR, '#ListingPageServicesSection div > a')
        for i in services_obj:
            provided_services.append(i.text)
        
        try:
            try:
                read_more = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#ListingPageAboutSection > div > .read-more-cont')))
                read_more.click()
                us = driver.find_elements(By.CSS_SELECTOR, '#ListingPageAboutSection > div > *')
                about_lst = ["".join(i.text) for i in us ]
                about_us  = ""
                for sent in about_lst:
                    #sentenc += " ".join(sent)
                    about_us = about_us+ " " + str(sent)
                about_us = about_us.replace('\n',' ')

            except: #if there is no read more button

                us = driver.find_elements(By.CSS_SELECTOR, '#ListingPageAboutSection > div > *')
                about_lst = ["".join(i.text) for i in us ]
                about_us  = ""
                for sent in about_lst:
                    #sentenc += " ".join(sent)
                    about_us = about_us+ " " + str(sent)
                about_us = about_us.replace('\n',' ')


        except:
            about_us = 'NA'



    
    elif service_data_type == 'basic':
        try:
            location =driver.find_element(By.CSS_SELECTOR, '.sidebar-box p').text
            facebook = 'NA'
            
            website = 'NA'
            about_us = 'NA'
            provided_services = 'NA'
        except:
            location = 'NA'
            facebook = 'NA'
            
            website = 'NA'
            about_us = 'NA'
            provided_services = 'NA'
    

    insta = 'NA'
    tktok = 'NA'
    phone_num = 'NA'
    email = 'NA'
    twt = 'NA'
    linkedin = 'NA'

    df = [serv_title, category_name, location,state, suburb,phone_num,website,email,facebook, insta,linkedin,twt,tktok,about_us,provided_services, data_type]
    return df


##### generating diagnostic file ########
########### function of generating excel file
def gen_file(category_name,state_name,suburb_name,folder_path,num_rows):
    
    file_path = folder_path

    status = 'Scrapped Successfully'
    scraped_rows= num_rows
    code_path = category_name+" "+state_name+" "+suburb_name
    if scraped_rows == 0 or file_path == 'NO DATA':
        file_path = 'NO DATA'
        status = 'NO DATA'
    
    
    scraped_time = now.strftime("%Y-%m-%d %H:%M:%S")
    try:
        ##read and write
        wb = load_workbook('Trade_service_diagnostic.xlsx')
        ws = wb.active
        
        ws.append([category_name, state_name, suburb_name,file_path,scraped_time,scraped_rows,status,code_path ])
        wb.save(filename='Trade_service_diagnostic.xlsx')
    except:
        headers  = ['category', 'state', 'suburb','file path','scrapped date','number of services', 'status', 'path code' ]
        workbook_name = 'Trade_service_diagnostic.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append([category_name, state_name, suburb_name,file_path,scraped_time,scraped_rows,status,code_path ])
        wb.save(workbook_name)
    return



###########  function of getting suburbs  ########
def scrap_each_suburb(suburb_link, category_name, state_name,suburb_name):

    
    # getting the services for each suburb
    # now we are in one suburb only
    services_lnks = serv_link_nd_type(suburb_link) #return services and types of each services
    try:
        df = pd.DataFrame(columns=["Business Name","Category" ,'Address' ,"State", "Suburb", "Contacts" ,"Website" ,"Email","Facebook", "Twitter", "Linkedin","Instagram", "Tiktok", "About us", "Types of Provided Services", "Data Type"])
    except:
        pass

    for j, serv in enumerate(services_lnks):
        #now we are in one service
        
        df_lsting = scrap(serv[0], category_name, state_name, suburb_name,serv[1]) #serv[0] is 'service link' and serv[1] is type of data for this service
        df.loc[len(df)+1] = df_lsting
    try:
        num_rows = len(df)
        folder_path = get_scoial_myfunc(df,category_name,state_name,suburb_name)
    except:
        num_rows = 0
        folder_path = 'NO DATA'
    
    gen_file(category_name,state_name,suburb_name,folder_path,num_rows)

    return num_rows

### function of sociale media scraping ######
def get_scoial_myfunc(df ,category_name,state_name,suburb_name):
    #lists for each link
    web_lst= []
    fb_lst =[]
    inst_lst= []
    tk_lst = []
    lnk_lst =[]
    twt_lst =[]
    email_miss_lst = []
    #get the website
    facebook_df = list(df['Facebook'])
    website = list(df['Website'])
    #print(len(website))
    for x,i in enumerate(website):
        #print(x)
        if str(i) == 'nan':
            #print("empty web")
            web_lst.append("NA")
            if pd.isna(df['Facebook'].iloc[x]) :
                
                fb_lst.append("NA")
                #print("ooh it's reaaaly here: ",str(facebook_df[x]))
                
            else:
                #print('here, the empty facebook is:', str(facebook_df[x]))
                fb_lst.append(df['Facebook'].iloc[x])

            if pd.isna(df['Instagram'].iloc[x]) :
                inst_lst.append("NA")
            else:
                inst_lst.append(df['Instagram'].iloc[x])

            
            tk_lst.append('NA')

            if pd.isna(df['Linkedin'].iloc[x]):
                lnk_lst.append('NA')
            else:
                lnk_lst.append(df['Linkedin'].iloc[x])
            
            if pd.isna(df['Twitter'].iloc[x]):
                twt_lst.append('NA')
            else:
                twt_lst.append(df['Twitter'].iloc[x])
            
        else:
            f = re.compile("https?://(www\.)?facebook\.com/(?!share\.php).(\S+\.?)+")  # regex for facebook
            matching = f.match(str(i))
            
            if bool(matching) is True:
                fb_lst.append(i)
                web_lst.append('NA')
                if pd.isna(df['Instagram'].iloc[x]) :
                    inst_lst.append("NA")
                else:
                    inst_lst.append(df['Instagram'].iloc[x])
                tk_lst.append('NA')

                if pd.isna(df['Linkedin'].iloc[x]):
                    lnk_lst.append('NA')
                else:
                    lnk_lst.append(df['Linkedin'].iloc[x])

                if pd.isna(df['Twitter'].iloc[x]):
                    twt_lst.append('NA')
                else:
                    twt_lst.append(df['Twitter'].iloc[x])


            else:
                website = i
                web_lst.append(website)
                
                for ss in range(3):
                    try:
                        html_data = None
                        check = None
                        if ss == 0:
                            try:
                                driver.get(str(website))
                                #print(str(website))
                                driver.set_page_load_timeout(10)
                                html_data = driver.page_source
                            except:
                                driver.get("https://"+str(website))
                                #print("https://"+str(website))
                                driver.set_page_load_timeout(10)
                                html_data = driver.page_source
                                
                    
                        bsoup = bs(html_data, 'html.parser')
                        #if bsoup is None:
                         #   print("the problem is here")
                        fb = bsoup.find('a', {'href': re.compile("https?://(www\.)?facebook\.com/(?!share\.php).(\S+\.?)+")})
                        if fb is None:
                            facebook = 'NA'
                            #print('cant find fb')
                        else:
                            facebook = fb['href']

                        #get email 
                        em = bsoup.find('a', {'href':re.compile(r"(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])")})
                        if em is None:   
                            ema = 'NA'
                        else:
                            ema = em['href']
                        
                        # get phone number r'((?:0|\+61|61)[234578]\d{8})'
                        ph = bsoup.find('a', {'href':re.compile(r"^(?:\+?(61))? ?(?:\((?=.*\)))?(0?[2-57-8])\)? ?(\d\d(?:[- ](?=\d{3})|(?!\d\d[- ]?\d[- ]))\d\d[- ]?\d[- ]?\d{3})|(13\s?(\d?\s?\d{3}?|\s?\d{2}\s?\d{2})|1[38]00\s?(\d{2}\s?\d{2}\s?\d{2}|\d{3}\s?\d{3}))$")})
                        if ph is None:
                            ph = 'NA'
                        else:
                            ph = ph['href']

                        ig = bsoup.find('a',
                                    {'href': re.compile(
                                        "https?://(www\.)?instagram\.com/(?!share\.php).(\S+\.?)+")})
                        if ig is None:
                            instagram = 'NA'
                        else:
                            instagram = ig['href']

                        tt = bsoup.find('a',
                                    {'href': re.compile(
                                    "https?://(www\.)?tiktok\.com/(?!share\.php).(\S+\.?)+")})
                        if tt is None:
                            tiktok = 'NA'
                        else:
                            tiktok = tt['href']

                        li = bsoup.find('a',
                                {'href': re.compile(
                                        "https?://(www\.)?linkedin\.com/(?!share\.php).(\S+\.?)+")})
                        if li is None:
                            linkedin = 'NA'
                        else:
                            linkedin = li['href']

                        tw = bsoup.find('a',
                                    {'href': re.compile(
                                        "https?://(www\.)?twitter\.com/(?!share\.php).(\S+\.?)+")})
                        if tw is None:
                            twitter = 'NA'
                        else:
                            twitter = tw['href']
                        break
                    except:
                        #print(f"Can't reach the website.\n{website}")
                        time.sleep(1)
                        facebook = 'NA'
                        instagram = 'NA'
                        tiktok = 'NA'
                        linkedin = 'NA'
                        twitter = 'NA'
                        continue
                
                if pd.isna(df['Facebook'].iloc[x]) :
                    fb_lst.append(facebook)   
                else:    
                    fb_lst.append(df['Facebook'].iloc[x])

                if pd.isna(df['Instagram'].iloc[x]) :
                    inst_lst.append(instagram)
                else:
                    inst_lst.append(df['Instagram'].iloc[x])


                if pd.isna(df['Linkedin'].iloc[x]):
                    lnk_lst.append(linkedin)
                else:
                    lnk_lst.append(df['Linkedin'].iloc[x])
                
                if pd.isna(df['Twitter'].iloc[x]):
                    twt_lst.append(twitter)
                else:
                    twt_lst.append(df['Twitter'].iloc[x])

                if pd.isna(df['Email'].iloc[x]):
                    email_miss_lst.append(ema)
                else:
                    email_miss_lst.append(df['Email'].iloc[x])
                    
                tk_lst.append(tiktok)
                
                
                                
            
    
 
    df['Website']=  web_lst
    df['Facebook']=  fb_lst
    df['Instagram']=  inst_lst
    df['Tiktok']=  tk_lst
    df['Linkedin']=  lnk_lst
    df['Twitter']=  twt_lst
    df['Email'] = get_email_facebook(df)
   

   #creating a folder for each state
    path = os.path.join(os.getcwd(),category_name+"/"+state_name)
    if not os.path.exists(path):
        os.makedirs(path)
      
        
        
    #pd.to_excel()
    try:
        #df.drop_duplicates(keep='first', inplace=True)
        df = df.replace(r'^\s*$', "NA", regex=True)
        df.to_excel(os.getcwd()+ "\\" +category_name+"\\"+state_name+"\\Trade_Service_"+category_name+"_"+state_name+"_"+suburb_name+".xlsx" , index=False,na_rep='NA')
    except:
        #df.drop_duplicates(keep='first', inplace=True)
        df = df.replace(r'^\s*$', "NA", regex=True)
        df.to_excel("Trade_Service_error.xlsx", index=False,na_rep='NA')
    folder_path = os.getcwd()+ "\\" +category_name+"\\"+state_name+"\\Trade_Service_"+category_name+"_"+state_name+"_"+suburb_name+".xlsx"
    return folder_path


#### function of driver get path ####
def get_path(path, usein, driver, ttl=30):
    WebDriverWait(driver, ttl).until(EC.presence_of_all_elements_located((usein, path)))
    return driver.find_element(by=usein, value=path)

######### start email from facebook function ###########
def get_email_facebook(df):

    facebook = list(df['Facebook'])
    email_lst = list(df['Email'])
    emails =[]

    for j , y in enumerate(email_lst):
        if pd.isna(df['Email'].iloc[j]) or df['Email'].iloc[j] == 'NA' or df['Email'].iloc[j] =='nan':
            fbok = df['Facebook'].iloc[j]
            if pd.isna(str(df['Facebook'].iloc[j])) or str(fbok) == 'nan' or str(fbok) == '' or str(fbok) =='NA':
                
                emails.append('NA')
                continue
            else:
                
                try:
                    driver.get(str(fbok))
                    
                    time.sleep(3)
                    email_rex = r'''(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])'''
                    Email = get_path("//body", By.XPATH, driver, 20).get_attribute('innerHTML')
                    final_email = re.search(email_rex, Email).group(0)
                    
                    if "login" in final_email or "profile" in final_email:
                        emails.append("NA")
                    else:
                        emails.append(final_email)
                except:
                    
                    final_email = "NA"
                    emails.append(final_email)
        else:
            emails.append(df['Email'].iloc[j])
            
            
    
    return emails

################### end email from facebook function #######







    



    ####### operating function #########
def operating(category_name):
    ## create folder for the category
    path = os.path.join(os.getcwd(),category_name)
    if not os.path.exists(path):
        os.makedirs(path)
    
    category_link = get_category_link(category_name) # this function returning category link



    states_lnks, states_names = get_states_links(category_link) #this function returning 8 links for each state and their names
    
    for i, st_lnk in enumerate(states_lnks): 
        try:    
            df_diagnostic = pd.read_excel('Trade_service_diagnostic.xlsx')
            state_latest_row = df_diagnostic['state'].loc[len(df_diagnostic['state'])-1]
            print()
            if states_names[i] != state_latest_row:
                continue
            if states_names[i] == state_latest_row:
                print(f"------------------------------------------ \
                \nnow you are scrpaing '{category_name}' in '{states_names[i]}' \
                \nwhich is the state number {i+1}/{len(states_names)} \n-----------------------------------")   

        except:
            print(f"------------------------------------------\
             \nnow you are scrpaing '{category_name}' in '{states_names[i]}' \
             \nwhich is the state number {i+1}/{len(states_names)} \n-----------------------------------")   
        
        suburbs_lnks, suburb_names = get_suburbs(st_lnk ) # this function takes 1 link from states links and returning suburbs links
    
        for j,x in enumerate(suburbs_lnks):
            try:
                df_diagnostic = pd.read_excel('Trade_service_diagnostic.xlsx')
                suburb_latest_row = df_diagnostic['suburb'].loc[len(df_diagnostic['suburb'])-1]
                if j == len(suburbs_lnks)-1  and suburb_names[len(suburbs_lnks)-1] == suburb_latest_row:
                    break
                
                if suburb_names[j] == suburb_latest_row:
                    print(f"--------------------------- \
                        \nnow you are scrapping {suburb_names[j]} \
                        \nwhich is the suburb number {j+1}/{len(suburbs_lnks)} \n-------------------------")
                    num_rows = scrap_each_suburb(x,category_name,states_names[i],suburb_names[j+1]) # x means one suburb
                else:
                    continue
                    
            except:
                print(f"--------------------------- \
                        \nnow you are scrapping {suburb_names[j]} \
                        \nwhich is the suburb number {j+1}/{len(suburbs_lnks)} \n-------------------------")
                num_rows = scrap_each_suburb(x,category_name,states_names[i],suburb_names[j]) # x means one suburb

            print(f"--------------------------- \
            \nyou scrapped {num_rows} from {suburb_names[j]} \
             \nwhich is the suburb number {j+1}/{len(suburbs_lnks)} \n-------------------------")
            



    #for st_lnk in states_lnks:
        #check the last states in the diagnostic file
        # then driver.get(st_lnk) which is equal to last state in the file
        # for suburb in suburbs : 
            #check the last suburb row if it is equal to the last suburb then continue to next state
            # else : scrap the next suburb  

    
    

    return print(f"------------------------------ congrats you finished scrapping the '{category_name}' category \
    \n------------------------------")

       
operating(category_name)
try:
    driver.dispose()
except:
    driver.quit()



